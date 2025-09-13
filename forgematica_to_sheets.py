#!/usr/bin/env python3

import argparse
import sys
import re
from pathlib import Path
from typing import Optional, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def guess_delimiter(sample_path: Path) -> str:
    # Try common delimiters; fall back to comma
    candidates = [",", ";", "\t", "|"]
    content = sample_path.read_text(errors="ignore")[:5000]
    counts = {d: content.count(d) for d in candidates}
    delimiter = max(counts, key=counts.get)
    return delimiter if counts[delimiter] > 0 else ","

def normalize_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.strip().lower())

def fuzzy_find(colnames: List[str], candidates: List[str]) -> Optional[str]:
    norm_to_orig = {normalize_header(c): c for c in colnames}
    # First pass: word boundary matches
    for orig in colnames:
        normed = normalize_header(orig)
        for cand in candidates:
            if re.search(rf"\b{re.escape(cand)}\b", normed):
                return orig
    # Second pass: substring matches
    for orig in colnames:
        normed = normalize_header(orig)
        for cand in candidates:
            if cand in normed:
                return orig
    return None

def detect_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    cols = list(df.columns)
    name_col = fuzzy_find(cols, ["name", "item", "material", "materials", "block", "id"])
    total_col = fuzzy_find(cols, ["total", "required", "qty_total", "quantity_total", "amount", "count"])
    missing_col = fuzzy_find(cols, ["missing", "needed", "to_get", "to_obtain", "short", "lack"])
    available_col = fuzzy_find(cols, ["available", "have", "stock", "in_chests", "present"])
    return {
        "name": name_col,
        "total": total_col,
        "missing": missing_col,
        "available": available_col
    }

def read_csv_any(path: Path, delimiter: Optional[str] = None) -> pd.DataFrame:
    if delimiter is None:
        delimiter = guess_delimiter(path)
    try:
        df = pd.read_csv(path, delimiter=delimiter)
    except Exception:
        df = pd.read_csv(path)  # try default
    return df

def build_base_table(df_grouped: pd.DataFrame, quantity_col: str,
                     include_user_cols: bool, default_stack_size: int) -> pd.DataFrame:
    out = pd.DataFrame()
    out["Materials"] = df_grouped["Materials"]
    out["Total (units)"] = df_grouped.get(quantity_col, pd.Series([0]*len(df_grouped)))
    if include_user_cols:
        out["User units (editable)"] = 0
        out["User stacks (editable)"] = 0
        out["Computed Total (units)"] = ""  # formula will fill
        out["Stack Size"] = ""             # formula / lookup
    else:
        out["Stack Size"] = default_stack_size
    # Common derived columns
    out["# Stacks (ceil)"] = ""
    out["# Double Chests"] = ""
    out["Stacks after last double"] = ""
    out["Units after last stack"] = ""
    return out

def write_df_with_formulas(ws, df_in: pd.DataFrame, is_missing_only: bool, default_stack_size: int):
    # Write header
    header_font = Font(bold=True)
    ws.append(list(df_in.columns))
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for _, row in df_in.iterrows():
        ws.append(list(row))

    # Map header names to column numbers
    headers = { ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1) }

    # Extract needed columns
    col_materials = headers.get("Materials")
    col_total = headers.get("Total (units)")
    col_stack = headers.get("Stack Size")
    col_stacksceil = headers.get("# Stacks (ceil)")
    col_double = headers.get("# Double Chests")
    col_stacks_after_dc = headers.get("Stacks after last double")
    col_units_after_stack = headers.get("Units after last stack")
    col_user_units = headers.get("User units (editable)")
    col_user_stacks = headers.get("User stacks (editable)")
    col_computed_total = headers.get("Computed Total (units)")

    # Data validation for non-negative inputs
    dv_nonneg = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws.add_data_validation(dv_nonneg)

    # For each data row (starting from row 2)
    for r in range(2, ws.max_row + 1):
        # If missing_only sheet: set Stack Size via VLOOKUP from REFS or default
        if is_missing_only and col_stack:
            # VLOOKUP(Materials, REFS!A:B, 2, FALSE), if fails, use default_stack_size
            ws.cell(row=r, column=col_stack, value=(
                f"=IFERROR(VLOOKUP({ws.cell(row=r, column=col_materials).coordinate}, REFS!A:B, 2, FALSE), {default_stack_size})"
            ))
        # Decide which total to use in downstream formulas
        if is_missing_only and col_computed_total and col_user_units and col_user_stacks and col_stack:
            # Computed Total = Missing (in Total (units) column) + user units + user stacks * stack size
            ws.cell(row=r, column=col_computed_total, value=(
                f"=MAX(0, {ws.cell(row=r, column=col_total).coordinate}"
                f"+{ws.cell(row=r, column=col_user_units).coordinate}"
                f"+{ws.cell(row=r, column=col_user_stacks).coordinate}*{ws.cell(row=r, column=col_stack).coordinate})"
            ))
            used_total_ref = ws.cell(row=r, column=col_computed_total).coordinate
        else:
            used_total_ref = ws.cell(row=r, column=col_total).coordinate

        # # Stacks (ceil)
        ws.cell(row=r, column=col_stacksceil, value=(
            f"=CEILING({used_total_ref}/{ws.cell(row=r, column=col_stack).coordinate}, 1)"
        ))

        # # Double Chests
        ws.cell(row=r, column=col_double, value=(
            f"=IF({used_total_ref}=0, 0, CEILING({ws.cell(row=r, column=col_stacksceil).coordinate}/54, 1))"
        ))

        # Stacks after last double
        ws.cell(row=r, column=col_stacks_after_dc, value=(
            f"=MOD({ws.cell(row=r, column=col_stacksceil).coordinate}, 54)"
        ))

        # Units after last stack
        ws.cell(row=r, column=col_units_after_stack, value=(
            f"=MOD({used_total_ref},{ws.cell(row=r, column=col_stack).coordinate})"
        ))

        # Apply data validation to editable fields
        for c in (col_stack, col_user_units, col_user_stacks):
            if c:
                dv_nonneg.add(ws.cell(row=r, column=c))

    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = 20

def generate_workbook(df_grouped: pd.DataFrame, out_path: Path, default_stack_size: int):
    wb = Workbook()
    ws_total = wb.active
    ws_total.title = "TOTALS_ALL"
    ws_missing = wb.create_sheet("MISSING_ONLY")
    ws_refs = wb.create_sheet("REFS")

    # TOTALS_ALL
    totals_all = build_base_table(df_grouped, "Total", include_user_cols=False, default_stack_size=default_stack_size)
    write_df_with_formulas(ws_total, totals_all, is_missing_only=False, default_stack_size=default_stack_size)

    # MISSING_ONLY
    missing_only = build_base_table(df_grouped, "Missing", include_user_cols=True, default_stack_size=default_stack_size)
    write_df_with_formulas(ws_missing, missing_only, is_missing_only=True, default_stack_size=default_stack_size)

    # REFS sheet: pre-fill some common materials & stack sizes + documentation
    ws_refs.append(["Materials", "Stack Size"])
    ws_refs["A1"].font = ws_refs["B1"].font = Font(bold=True)
    common = [
        ["Ender Pearl", 16],
        ["Egg", 16],
        ["Snowball", 16],
        ["Boat", 1],
        ["Armor (any)", 1],
        ["Tool (any)", 1],
        ["Banner", 16],
        # etc — you can edit or expand this list
    ]
    for row in common:
        ws_refs.append(row)

    # Add a blank row, then notes / links
    ws_refs.append([])
    ws_refs.append(["Docs", "URL"])
    ws_refs["A"+str(ws_refs.max_row)].font = ws_refs["B"+str(ws_refs.max_row)].font = Font(bold=True)
    docs = [
        ("Google Sheets – CEILING", "https://support.google.com/docs/answer/3093471"),
        ("Google Sheets – MOD", "https://support.google.com/docs/answer/3093497"),
        ("Google Sheets – VLOOKUP", "https://support.google.com/docs/answer/3093318"),
        ("Minecraft – Double Chest (54 slots)", "https://minecraft.fandom.com/wiki/Chest"),
    ]
    for label, url in docs:
        ws_refs.append([label, url])

    # Save workbook
    wb.save(str(out_path))

def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Convert Forgematica-like CSV into Sheets-ready .xlsx with formulas")
    parser.add_argument("--csv", required=True, help="Path to input CSV file")
    parser.add_argument("--out", default="forgematica_materials_sheets.xlsx", help="Path to output .xlsx")
    parser.add_argument("--delimiter", default=None, help="Delimiter for CSV (auto-detect if omitted)")
    parser.add_argument("--name-col", default=None, help="Override name/material column")
    parser.add_argument("--total-col", default=None, help="Override total/required column")
    parser.add_argument("--missing-col", default=None, help="Override missing/needed column")
    parser.add_argument("--available-col", default=None, help="Override available/have column")
    parser.add_argument("--default-stack-size", type=int, default=64, help="Default stack size if no lookup matches")
    args = parser.parse_args(argv)

    csv_path = Path(args.csv)
    if not csv_path.is_file():
        print(f"Error: CSV file not found: {csv_path}", file=sys.stderr)
        return 1

    df_raw = read_csv_any(csv_path, delimiter=args.delimiter)

    # Detect or override columns
    mapping = detect_columns(df_raw)
    if args.name_col:
        mapping["name"] = args.name_col
    if args.total_col:
        mapping["total"] = args.total_col
    if args.missing_col:
        mapping["missing"] = args.missing_col
    if args.available_col:
        mapping["available"] = args.available_col

    # Build a DataFrame with renamed canonical cols
    rename_map = {}
    if mapping["name"] and mapping["name"] in df_raw.columns:
        rename_map[mapping["name"]] = "Materials"
    else:
        # fallback
        df_raw["__Materials__"] = "Unknown"
        rename_map["__Materials__"] = "Materials"

    if mapping["total"] and mapping["total"] in df_raw.columns:
        rename_map[mapping["total"]] = "Total"
    if mapping["missing"] and mapping["missing"] in df_raw.columns:
        rename_map[mapping["missing"]] = "Missing"
    if mapping["available"] and mapping["available"] in df_raw.columns:
        rename_map[mapping["available"]] = "Available"

    df = df_raw[list(rename_map.keys())].rename(columns=rename_map).copy()

    # Convert numeric columns
    for c in ("Total", "Missing", "Available"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)

    # Group by Materials
    agg = {}
    if "Total" in df.columns:
        agg["Total"] = "sum"
    if "Missing" in df.columns:
        agg["Missing"] = "sum"
    if "Available" in df.columns:
        agg["Available"] = "sum"

    if agg:
        df_grouped = df.groupby("Materials", as_index=False).agg(agg)
    else:
        # minimal fallback
        df_grouped = df.drop_duplicates("Materials").copy()
        for c in ("Total", "Missing", "Available"):
            if c not in df_grouped.columns:
                df_grouped[c] = 0

    out_path = Path(args.out)
    generate_workbook(df_grouped, out_path, default_stack_size=args.default_stack_size)
    print(f"Workbook written to {out_path.resolve()}")

    return 0

if __name__ == "__main__":
    sys.exit(main())
